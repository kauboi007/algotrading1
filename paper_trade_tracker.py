import yfinance as yf
import pandas as pd
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TRACKER_FILE = "paper_trade_log.xlsx"
STATE_FILE   = "paper_trade_state.json"
CAPITAL      = 100000  # 1L INR
COST_RATE    = 0.0022  # 0.22% round trip Zerodha delivery

# ── colours ──────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BUY_FILL  = PatternFill("solid", start_color="C6EFCE")
SELL_FILL = PatternFill("solid", start_color="FFC7CE")
HOLD_FILL = PatternFill("solid", start_color="DDEBF7")
BEAR_FILL = PatternFill("solid", start_color="FFF2CC")
ALT_FILL  = PatternFill("solid", start_color="F2F2F2")

def _font(bold=False, size=10):
    return Font(name="Arial", size=size, bold=bold)
def _ctr():
    return Alignment(horizontal="center", vertical="center")

# ── state ─────────────────────────────────────────────────
def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            return json.load(f)
    return {
        "filtered":   {"months": [], "prev_weights": {}},
        "unfiltered": {"months": [], "prev_weights": {}}
    }

def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2, default=str)

# ── prices ────────────────────────────────────────────────
def get_prices(tickers):
    prices = {}

    for t in tickers:
        try:
            d = yf.download(
                t,
                period="15d",
                auto_adjust=False,
                progress=False
            )

            if d.empty:
                prices[t] = None
                continue

            close_series = d["Close"].dropna()

            if close_series.empty:
                prices[t] = None
                continue

            # latest available trading day's close
            prices[t] = round(float(close_series.squeeze().iloc[-1]), 2)

        except Exception:
            prices[t] = None

    return prices


def get_nifty_price():
    try:
        d = yf.download(
            "^CRSLDX",
            period="15d",
            auto_adjust=False,
            progress=False
        )

        if d.empty:
            return None

        close_series = d["Close"].dropna()

        if close_series.empty:
            return None

        # latest available trading day's close
        return round(float(close_series.squeeze().iloc[-1]), 2)

    except Exception:
        return None
# ── Excel init ────────────────────────────────────────────
TRADE_HEADERS = [
    "Month", "Ticker", "Action", "Weight %", "Amount (₹)",
    "Shares", "Buy Price (₹)", "Sell Price (₹)", "P&L (₹)", "Return %", "Notes"
]
TRADE_WIDTHS = [12, 18, 8, 10, 12, 8, 16, 16, 12, 10, 20]

SUM_HEADERS = [
    "Month", "Regime", "Capital (₹)", "# Stocks", "Turnover",
    "Cost (₹)", "Actual Return %", "NIFTY 500 %", "Alpha %", "Notes"
]
SUM_WIDTHS = [12, 12, 14, 10, 10, 10, 16, 14, 12, 25]

def _make_sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    ws.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = _ctr()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    return ws

def init_excel():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _make_sheet(wb, "Summary (Filtered)",   SUM_HEADERS,   SUM_WIDTHS)
    _make_sheet(wb, "Summary (Unfiltered)", SUM_HEADERS,   SUM_WIDTHS)
    _make_sheet(wb, "Trades (Filtered)",    TRADE_HEADERS, TRADE_WIDTHS)
    _make_sheet(wb, "Trades (Unfiltered)",  TRADE_HEADERS, TRADE_WIDTHS)

    wb.save(TRACKER_FILE)
    print(f"✅ Created {TRACKER_FILE}")

# ── core logic ────────────────────────────────────────────
def _compute_trades(weights_dict, prev_weights, nifty_price):
    all_stocks = set(weights_dict) | set(prev_weights)
    turnover   = sum(abs(weights_dict.get(s,0) - prev_weights.get(s,0)) for s in all_stocks)
    cost       = round(turnover * COST_RATE * CAPITAL, 2)
    deployable = CAPITAL - cost

    actions = {}
    for s in all_stocks:
        if s in weights_dict and s not in prev_weights:   actions[s] = "BUY"
        elif s not in weights_dict and s in prev_weights: actions[s] = "SELL"
        else:                                              actions[s] = "HOLD"

    tickers_needed = list(weights_dict.keys())
    print("📡 Fetching buy prices …")
    buy_prices = get_prices(tickers_needed)

    trades = []
    for ticker, weight in weights_dict.items():
        amount = round(weight * deployable, 2)
        bp     = buy_prices.get(ticker)
        shares = int(amount / bp) if bp else 0
        trades.append({
            "ticker": ticker, "action": actions.get(ticker, "BUY"),
            "weight": weight, "amount": amount,
            "buy_price": bp, "shares": shares,
            "sell_price": None, "pnl": None
        })
    for ticker in prev_weights:
        if ticker not in weights_dict:
            trades.append({
                "ticker": ticker, "action": "SELL",
                "weight": 0, "amount": 0,
                "buy_price": None, "shares": 0,
                "sell_price": None, "pnl": None
            })
    return trades, turnover, cost, deployable

def _write_trades_to_sheet(wb, sheet_name, month_label, trades):
    wt       = wb[sheet_name]
    next_row = wt.max_row + 1
    action_fill = {"BUY": BUY_FILL, "SELL": SELL_FILL, "HOLD": HOLD_FILL}
    for i, t in enumerate(trades):
        row  = next_row + i
        fill = action_fill.get(t["action"], ALT_FILL)
        vals = [
            month_label, t["ticker"], t["action"],
            f"{t['weight']*100:.2f}%" if t["weight"] else "—",
            t["amount"] if t["amount"] else "—",
            t["shares"] if t["shares"] else "—",
            t["buy_price"] if t["buy_price"] else "TBC",
            "", "", "", ""
        ]
        for col, val in enumerate(vals, 1):
            c = wt.cell(row=row, column=col, value=val)
            c.font = _font(); c.fill = fill; c.alignment = _ctr()

def _write_summary_to_sheet(wb, sheet_name, month_label, regime,
                             n_stocks, turnover, cost,
                             actual_return, nifty_return, alpha, note):
    ws   = wb[sheet_name]
    nr   = ws.max_row + 1
    fill = BUY_FILL if (alpha or 0) >= 0 else SELL_FILL
    vals = [
        month_label, regime, CAPITAL, n_stocks,
        f"{turnover:.2%}", cost,
        f"{actual_return:.2f}%",
        f"{nifty_return:.2f}%" if nifty_return is not None else "N/A",
        f"{alpha:.2f}%"        if alpha        is not None else "N/A",
        note
    ]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=nr, column=col, value=val)
        c.font = _font(); c.alignment = _ctr()
        if col == 9: c.fill = fill

def _print_trades(month_label, trades, cost, turnover, deployable):
    print(f"\n{'='*65}")
    print(f"  MONTH: {month_label}  |  Capital: ₹{CAPITAL:,}  |  Cost: ₹{cost:.2f}")
    print(f"  Turnover: {turnover:.2%}  |  Deployable: ₹{deployable:,.2f}")
    print(f"{'='*65}")
    print(f"  {'Ticker':<20} {'Action':<6} {'Weight':>8} {'Amount':>12} {'Price':>10} {'Shares':>7}")
    print(f"  {'-'*63}")
    for t in trades:
        print(f"  {t['ticker']:<20} {t['action']:<6} "
              f"{t['weight']*100:>7.2f}%  "
              f"{str(t['amount']):>12}  "
              f"{str(t['buy_price']):>10}  "
              f"{t['shares']:>7}")

# ── PUBLIC API ────────────────────────────────────────────

def start_month(weights_dict, is_bull, month_label=None):
    """
    Call at end of month. Logs BOTH filtered and unfiltered trades.
    is_bull: bool from regime_detection()
    """
    state = load_state()
    if not os.path.exists(TRACKER_FILE):
        init_excel()

    if month_label is None:
        nxt = pd.Timestamp.today() + pd.offsets.MonthBegin(1)
        month_label = nxt.strftime("%b-%Y")

    nifty_price = get_nifty_price()

    # ── UNFILTERED (always runs) ──────────────────────────
    uf_state = state["unfiltered"]
    already  = any(m["month"] == month_label for m in uf_state["months"])
    if not already:
        trades, turnover, cost, deployable = _compute_trades(
            weights_dict, uf_state["prev_weights"], nifty_price)

        uf_state["months"].append({
            "month": month_label, "regime": "Bull" if is_bull else "Bear*",
            "capital": CAPITAL, "cost": cost,
            "turnover": round(turnover,4), "trades": trades,
            "nifty_buy": nifty_price, "status": "open"
        })
        uf_state["prev_weights"] = weights_dict

        wb = load_workbook(TRACKER_FILE)
        _write_trades_to_sheet(wb, "Trades (Unfiltered)", month_label, trades)
        wb.save(TRACKER_FILE)

        print(f"\n🟡 UNFILTERED — {month_label}")
        _print_trades(month_label, trades, cost, turnover, deployable)

    # ── FILTERED (only if bull) ───────────────────────────
    f_state = state["filtered"]
    already = any(m["month"] == month_label for m in f_state["months"])
    if not already:
        if is_bull:
            trades_f, turnover_f, cost_f, deployable_f = _compute_trades(
                weights_dict, f_state["prev_weights"], nifty_price)

            f_state["months"].append({
                "month": month_label, "regime": "Bull",
                "capital": CAPITAL, "cost": cost_f,
                "turnover": round(turnover_f,4), "trades": trades_f,
                "nifty_buy": nifty_price, "status": "open"
            })
            f_state["prev_weights"] = weights_dict

            wb = load_workbook(TRACKER_FILE)
            _write_trades_to_sheet(wb, "Trades (Filtered)", month_label, trades_f)
            wb.save(TRACKER_FILE)

            print(f"\n🟢 FILTERED — {month_label} (Bull — investing)")
            _print_trades(month_label, trades_f, cost_f, turnover_f, deployable_f)
        else:
            # bear month — log as cash in filtered sheet
            f_state["months"].append({
                "month": month_label, "regime": "Bear",
                "capital": CAPITAL, "cost": 0, "turnover": 0,
                "trades": [], "nifty_buy": nifty_price,
                "status": "closed", "actual_return": 0,
                "nifty_return": None, "alpha": None, "total_pnl": 0
            })
            wb = load_workbook(TRACKER_FILE)
            ws = wb["Summary (Filtered)"]
            nr = ws.max_row + 1
            vals = [month_label, "Bear (Cash)", CAPITAL, 0,
                    "0%", 0, "0%", "N/A", "N/A", "Death cross — sat in cash"]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=nr, column=col, value=val)
                c.font = _font(); c.fill = BEAR_FILL; c.alignment = _ctr()
            wb.save(TRACKER_FILE)
            print(f"\n🐻 FILTERED — {month_label} (Bear — cash)")

    save_state(state)
    print(f"\n✅ Logged to {TRACKER_FILE}")


def end_month(month_label=None):
    """Call after market close on last trading day of the month."""
    state = load_state()
    if month_label is None:
        month_label = pd.Timestamp.today().strftime("%b-%Y")

    nifty_sell = get_nifty_price()
    wb = load_workbook(TRACKER_FILE)

    for strategy in ["filtered", "unfiltered"]:
        s         = state[strategy]
        month_data = next((m for m in s["months"] if m["month"] == month_label), None)
        if not month_data or month_data["status"] == "closed":
            continue
        if not month_data["trades"]:
            continue

        active = [t["ticker"] for t in month_data["trades"] if t["action"] != "SELL"]
        print(f"📡 Fetching sell prices ({strategy}) …")
        sell_prices = get_prices(active)

        total_pnl = 0
        for t in month_data["trades"]:
            if t["action"] == "SELL" or not t["buy_price"] or not t["shares"]:
                t["sell_price"] = None; t["pnl"] = 0; continue
            sp = sell_prices.get(t["ticker"])
            t["sell_price"] = sp
            if sp:
                pnl = (sp - t["buy_price"]) * t["shares"]
                t["pnl"] = round(pnl, 2); total_pnl += pnl

        actual_return = round((total_pnl / CAPITAL) * 100, 4)
        nifty_return  = None
        if nifty_sell and month_data.get("nifty_buy"):
            nifty_return = round(
                ((nifty_sell - month_data["nifty_buy"]) / month_data["nifty_buy"]) * 100, 4)
        alpha = round(actual_return - nifty_return, 4) if nifty_return is not None else None

        month_data.update({
            "status": "closed", "actual_return": actual_return,
            "nifty_return": nifty_return, "alpha": alpha,
            "total_pnl": round(total_pnl, 2)
        })

        # update trades sheet
        sheet_name = f"Trades ({'Filtered' if strategy == 'filtered' else 'Unfiltered'})"
        wt = wb[sheet_name]
        for row in range(2, wt.max_row + 1):
            if wt.cell(row=row, column=1).value != month_label: continue
            ticker = wt.cell(row=row, column=2).value
            trade  = next((t for t in month_data["trades"] if t["ticker"] == ticker), None)
            if not trade: continue
            sp  = trade.get("sell_price")
            bp  = trade.get("buy_price")
            pnl = trade.get("pnl", 0)
            ret = round(((sp-bp)/bp)*100, 2) if sp and bp else None
            wt.cell(row=row, column=8).value  = sp  if sp  else "—"
            wt.cell(row=row, column=9).value  = pnl if pnl else "—"
            wt.cell(row=row, column=10).value = f"{ret:.2f}%" if ret is not None else "—"

        # write summary
        sum_name = f"Summary ({'Filtered' if strategy == 'filtered' else 'Unfiltered'})"
        _write_summary_to_sheet(
            wb, sum_name, month_label,
            month_data["regime"],
            len([t for t in month_data["trades"] if t["action"] != "SELL"]),
            month_data["turnover"], month_data["cost"],
            actual_return, nifty_return, alpha,
            f"P&L: ₹{total_pnl:,.2f}"
        )

        print(f"\n{'='*55}")
        print(f"  [{strategy.upper()}] {month_label} CLOSED")
        print(f"  Return: {actual_return:.2f}%  |  "
              f"NIFTY: {nifty_return:.2f}%  |  Alpha: {alpha:.2f}%")
        print(f"  Total P&L: ₹{total_pnl:,.2f}")

    wb.save(TRACKER_FILE)
    save_state(state)
    print(f"\n✅ {TRACKER_FILE} updated")


def show_status():
    state = load_state()
    for strategy in ["filtered", "unfiltered"]:
        s = state[strategy]
        print(f"\n{'='*55}")
        print(f"  {strategy.upper()}  |  Capital: ₹{CAPITAL:,}")
        print(f"{'='*55}")
        total_pnl = 0
        for m in s["months"]:
            ret   = m.get("actual_return", 0) or 0
            alpha = m.get("alpha")
            pnl   = m.get("total_pnl", 0) or 0
            total_pnl += pnl
            a_str = f"  α={alpha:+.2f}%" if alpha is not None else ""
            print(f"  {m['month']:<12} {m['regime']:<12} "
                  f"{m['status']:<8} ret={ret:+.2f}%{a_str}")
        print(f"  Cumulative P&L: ₹{total_pnl:,.2f}")